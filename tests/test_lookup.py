from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from lp_lookup.adapters import ExcelSourceAdapter
from lp_lookup.matching import InvestorMatcher
from lp_lookup.service import LookupService


def _write_workbooks(tmp_path: Path) -> tuple[Path, Path]:
    lp_path = tmp_path / "lp.xlsx"
    company_path = tmp_path / "company.xlsx"

    lp_wb = Workbook()
    ws_pairs = lp_wb.active
    ws_pairs.title = "All LP-Fund Pairs"
    ws_pairs.append(
        [
            "Fund Manager",
            "LP Name",
            "LP Type",
            "City",
            "Country",
            "# Funds with Manager",
            "Source",
            "Scraped Date",
        ]
    )
    ws_pairs.append(["Accel", "LP Alpha", "Endowment", "London", "UK", 1, "src", "today"])
    ws_pairs.append(["Alpha Wave Global", "LP Alpha", "Endowment", "London", "UK", 1, "src", "today"])
    ws_pairs.append(["Alpha Wave Global", "LP Beta", "Pension", "New York", "US", 1, "src", "today"])
    ws_pairs.append(["Andreessen Horowitz", "LP Gamma", "Foundation", "San Francisco", "US", 1, "src", "today"])

    ws_unique = lp_wb.create_sheet("Unique LPs")
    ws_unique.append(["LP Name", "LP Type", "City", "Country", "# Fund Managers", "Fund Manager Names"])
    ws_unique.append(["LP Alpha", "Endowment", "London", "UK", 2, "Accel, Alpha Wave Global"])
    ws_unique.append(["LP Beta", "Pension", "New York", "US", 1, "Alpha Wave Global"])
    ws_unique.append(["LP Gamma", "Foundation", "San Francisco", "US", 1, "Andreessen Horowitz"])

    lp_wb.save(lp_path)

    company_wb = Workbook()
    ws_company = company_wb.active
    ws_company.title = "Company-Fund Map"
    ws_company.append(["Company", "Lead Investors"])
    ws_company.append(["Anthropic", "Accel, Alpha Wave Ventures, Unknown Capital"])
    ws_company.append(["World Labs", "Andreessen Horowitz"])
    company_wb.save(company_path)

    return lp_path, company_path


def test_excel_adapter_splits_investors(tmp_path: Path) -> None:
    lp_path, company_path = _write_workbooks(tmp_path)
    adapter = ExcelSourceAdapter(lp_database_path=lp_path, company_lookup_path=company_path)

    company_rows = list(adapter.load_company_investors())

    assert [row.raw_investor for row in company_rows if row.company == "Anthropic"] == [
        "Accel",
        "Alpha Wave Ventures",
        "Unknown Capital",
    ]


def test_matcher_handles_exact_normalized_and_thresholded_fuzzy() -> None:
    matcher = InvestorMatcher(
        [
            "Accel",
            "Alpha Wave Global",
            "Andreessen Horowitz",
            "Monk's Hill Ventures",
            "Greylock Partners",
            "PremjiInvest",
            "Founders Fund",
            "Advanced Technology Ventures",
        ]
    )

    exact = matcher.match(company="Anthropic", raw_investor="Accel")
    assert [item.match_method for item in exact.matches] == ["exact"]

    normalized = matcher.match(company="Anthropic", raw_investor="Monks Hill Ventures")
    assert [item.match_method for item in normalized.matches] == ["normalized"]

    fuzzy = matcher.match(company="Anthropic", raw_investor="Alpha Wave Ventures")
    assert fuzzy.matches
    assert all(item.match_method == "fuzzy" for item in fuzzy.matches)
    assert all(item.match_score >= 70 for item in fuzzy.matches)

    good_fuzzy = matcher.match(company="Anthropic", raw_investor="Greylock")
    assert [item.matched_fund_manager for item in good_fuzzy.matches] == ["Greylock Partners"]

    compact_fuzzy = matcher.match(company="Anthropic", raw_investor="Premji Invest")
    assert [item.matched_fund_manager for item in compact_fuzzy.matches] == ["PremjiInvest"]

    bad_suffix_fuzzy = matcher.match(company="Anthropic", raw_investor="Ballistic Ventures")
    assert bad_suffix_fuzzy.matches == ()
    assert bad_suffix_fuzzy.best_candidate is not None

    bad_fund_fuzzy = matcher.match(company="Anthropic", raw_investor="CrowdStrike Falcon Fund")
    assert bad_fund_fuzzy.matches == ()
    assert bad_fund_fuzzy.best_candidate is not None

    unmatched = matcher.match(company="Anthropic", raw_investor="Totally Different Investor")
    assert unmatched.matches == ()
    assert unmatched.best_score is not None
    assert unmatched.best_score < 70


def test_lookup_service_dedupes_lps_and_tracks_unmatched(tmp_path: Path) -> None:
    lp_path, company_path = _write_workbooks(tmp_path)
    service = LookupService(ExcelSourceAdapter(lp_database_path=lp_path, company_lookup_path=company_path))

    exposure_rows = service.get_exposure_rows("Anthropic")
    exposure_by_lp = {row.lp_name: row for row in exposure_rows}

    assert set(exposure_by_lp) == {"LP Alpha", "LP Beta"}
    assert exposure_by_lp["LP Alpha"].matched_investors == ("Accel", "Alpha Wave Global")
    assert exposure_by_lp["LP Alpha"].best_score == 100

    unmatched_rows = service.get_unmatched_rows("Anthropic")
    assert [row.raw_investor for row in unmatched_rows] == ["Unknown Capital"]

    summary = service.company_summary("Anthropic")
    assert summary == {
        "matched_investors": 2,
        "unmatched_investors": 1,
        "deduped_lps": 2,
    }


@pytest.mark.skipif(
    not Path("/Users/pascalsuhrcke/Downloads/Atrea_LP_Database_Export.xlsx").exists()
    or not Path("/Users/pascalsuhrcke/Downloads/Company Look-Up.xlsx").exists(),
    reason="Real source workbooks are not available",
)
@pytest.mark.parametrize("company", ["Anthropic", "World Labs", "Glean"])
def test_real_workbook_smoke(company: str) -> None:
    service = LookupService()
    assert company in service.list_companies()
    summary = service.company_summary(company)
    assert summary["matched_investors"] >= 1
