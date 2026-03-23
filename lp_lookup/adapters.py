from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .config import (
    COMPANY_FUND_MAP_SHEET,
    COMPANY_LOOKUP_HEADERS,
    COMPANY_LOOKUP_PATH,
    LP_DATABASE_PATH,
    LP_FUND_PAIR_HEADERS,
    LP_FUND_PAIR_SHEET,
    UNIQUE_LP_HEADERS,
    UNIQUE_LPS_SHEET,
)
from .models import CompanyInvestor, LPFundPair, LPRecord, SourceAdapter


class WorkbookValidationError(ValueError):
    pass


def _clean_cell(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _require_headers(actual: tuple[str, ...], expected: tuple[str, ...], sheet_name: str) -> None:
    if actual[: len(expected)] != expected:
        raise WorkbookValidationError(
            f"Unexpected headers in sheet '{sheet_name}'. "
            f"Expected {expected}, found {actual[:len(expected)]}."
        )


class ExcelSourceAdapter(SourceAdapter):
    def __init__(
        self,
        lp_database_path: Path = LP_DATABASE_PATH,
        company_lookup_path: Path = COMPANY_LOOKUP_PATH,
    ) -> None:
        self.lp_database_path = Path(lp_database_path)
        self.company_lookup_path = Path(company_lookup_path)

    @property
    def source_paths(self) -> tuple[Path, ...]:
        return (self.lp_database_path, self.company_lookup_path)

    def _load_workbook(self, path: Path):
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")
        return load_workbook(path, read_only=True, data_only=True)

    def load_lp_fund_pairs(self) -> Iterable[LPFundPair]:
        workbook = self._load_workbook(self.lp_database_path)
        sheet = workbook[LP_FUND_PAIR_SHEET]
        header = tuple(_clean_cell(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        _require_headers(header, LP_FUND_PAIR_HEADERS, LP_FUND_PAIR_SHEET)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            fund_manager = _clean_cell(row[0])
            lp_name = _clean_cell(row[1])
            if not fund_manager or not lp_name:
                continue
            yield LPFundPair(
                fund_manager=fund_manager,
                lp_name=lp_name,
                lp_type=_clean_cell(row[2]),
                city=_clean_cell(row[3]),
                country=_clean_cell(row[4]),
            )

    def load_unique_lps(self) -> Iterable[LPRecord]:
        workbook = self._load_workbook(self.lp_database_path)
        sheet = workbook[UNIQUE_LPS_SHEET]
        header = tuple(_clean_cell(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        _require_headers(header, UNIQUE_LP_HEADERS, UNIQUE_LPS_SHEET)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            lp_name = _clean_cell(row[0])
            if not lp_name:
                continue
            yield LPRecord(
                lp_name=lp_name,
                lp_type=_clean_cell(row[1]),
                city=_clean_cell(row[2]),
                country=_clean_cell(row[3]),
            )

    def load_company_investors(self) -> Iterable[CompanyInvestor]:
        workbook = self._load_workbook(self.company_lookup_path)
        sheet = workbook[COMPANY_FUND_MAP_SHEET]
        header = tuple(_clean_cell(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        _require_headers(header, COMPANY_LOOKUP_HEADERS, COMPANY_FUND_MAP_SHEET)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            company = _clean_cell(row[0])
            investor_blob = _clean_cell(row[1])
            if not company:
                continue
            if not investor_blob:
                continue
            for raw_investor in (part.strip() for part in investor_blob.split(",")):
                if raw_investor:
                    yield CompanyInvestor(company=company, raw_investor=raw_investor)
